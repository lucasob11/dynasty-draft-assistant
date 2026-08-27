#!/usr/bin/env python3
"""Rebuilds rankings.json from the dynasty draft prep workbook:
  - overall rank/team/bye/age from 'Footballers 2QB (Full 472)'
  - per-position tier from the QB/RB/WR/TE 'Tier List' tabs (tiers are
    delimited by blank rows)
  - multi-source expert consensus (best/worst/avg rank, and per-source
    ranks for head-to-head comparisons) from 'Sheet1', which holds six
    experts' top-20-per-position boards: Trade/Cut Rankings, FantasyPros,
    ANDY/JASON/MIKE (FantasyFootballers), and DraftSharks
  - 2025 season stats (games played, PPR points/game, season-end position
    rank) from a Pro-Football-Reference fantasy leaders export — a
    same-origin, no-JS "xls" download that's actually an HTML table

Any player who appears in Sheet1's top 20 but isn't already in the 472-list
gets appended as a new entry (tier/bye/age left blank — nothing else in the
workbook has that for them) so their expert data isn't silently dropped.

Re-run this any time the spreadsheet or the stats export is updated:

    python3 extract_rankings.py ~/Downloads/2026_Dynasty_Startup_Draft_Prep.xlsx ~/Downloads/sportsref_download.xls

Both arguments are optional and default to those exact filenames in ~/Downloads.
If the stats file isn't found, 2025 stats are just skipped (not a hard failure).
"""
import difflib
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

RANK_SHEET = "Footballers 2QB (Full 472)"
TIER_SHEETS = {
    "QB": "QB Tier List",
    "RB": "RB Tier List",
    "WR": "WR Tier List",
    "TE": "TE Tier List",
}
VALID_POS = set(TIER_SHEETS)

EXPERT_SHEET = "Sheet1"
EXPERT_SOURCES = [
    ("Trade/Cut", 2),    # columns B-E, QB/RB/WR/TE — abbreviated names, no team
    ("FantasyPros", 7),  # columns G-J — "Name (TEAM)"
    ("ANDY", 12),        # columns L-O
    ("JASON", 17),       # columns Q-T
    ("MIKE", 22),        # columns V-Y
    ("DraftSharks", 27), # columns AA-AD
]
EXPERT_POSITIONS = ["QB", "RB", "WR", "TE"]
EXPERT_ROWS = range(4, 24)  # ranks 1-20

# Trade/Cut Rankings uses ad-hoc abbreviations/nicknames that the normal
# last-name/first-name/substring/fuzzy resolution in resolve_abbrev() can't
# confidently place. Verified by hand against the sheet.
MANUAL_NAME_OVERRIDES = {
    ("RB", "cmc"): "Christian McCaffrey",
    ("RB", "treyveon"): "TreVeyon Henderson",
    ("WR", "jefferon"): "Justin Jefferson",
}

SUFFIX_RE = re.compile(r"\b(jr|sr|ii|iii|iv|v)\b")
TEAM_SUFFIX_RE = re.compile(r"\s*\(([A-Za-z]{2,3})\)\s*$")


def normalize_name(name):
    if not name:
        return ""
    n = unicodedata.normalize("NFD", str(name)).encode("ascii", "ignore").decode()
    n = n.lower()
    n = n.replace(".", "").replace("'", "").replace("’", "")
    n = n.replace("-", " ")
    n = SUFFIX_RE.sub("", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def last_name(normalized):
    parts = normalized.split(" ")
    return parts[-1] if parts else ""


def load_tiers(wb):
    """pos -> {normalized_name: (tier_int, pos_rank)}, plus a last-name fallback map.

    Tiers are delimited by blank rows in each Tier List sheet. pos_rank (column A)
    is captured too because the sheet's row order isn't always sorted by it (a
    couple of rows are manually out of order) — sorting by the numeric pos_rank
    is what actually keeps each tier's players in a stable, correct order.
    """
    exact = {}
    by_last = {}  # (pos, last) -> set of (tier, pos_rank) seen
    for pos, sheet_name in TIER_SHEETS.items():
        ws = wb[sheet_name]
        tier = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            pos_rank, name = row[0], row[1]
            if name is None:
                tier += 1
                continue
            if tier == 0:
                tier = 1
            key = normalize_name(name)
            if not key:
                continue
            entry = (tier, float(pos_rank) if pos_rank is not None else 9999)
            exact[(pos, key)] = entry
            by_last.setdefault((pos, last_name(key)), set()).add(entry)
    return exact, by_last


def resolve_abbrev(pos, token, candidates):
    """Resolves a Trade/Cut abbreviation/nickname to a display name from
    `candidates` (normalized_full -> display_name, built from the five
    full-name sources for this position). Returns None rather than guessing
    when nothing lines up unambiguously."""
    nt = normalize_name(token)
    override = MANUAL_NAME_OVERRIDES.get((pos, nt))
    if override:
        return override
    if nt in candidates:
        return candidates[nt]
    words = nt.split()
    if not words:
        return None
    last = words[-1]
    last_matches = [d for n, d in candidates.items() if n.split()[-1] == last]
    if len(last_matches) == 1:
        return last_matches[0]
    if len(words) == 1:
        first_matches = [d for n, d in candidates.items() if n.split()[0] == words[0]]
        if len(first_matches) == 1:
            return first_matches[0]
    sub_matches = [d for n, d in candidates.items() if nt in n or n in nt]
    if len(sub_matches) == 1:
        return sub_matches[0]
    close = difflib.get_close_matches(nt, list(candidates.keys()), n=1, cutoff=0.72)
    if close:
        return candidates[close[0]]
    return None


def load_expert_consensus(wb):
    """pos -> {normalized_name: {"display", "team", "ranks": {source: rank}}}"""
    if EXPERT_SHEET not in wb.sheetnames:
        return {}
    ws = wb[EXPERT_SHEET]

    raw = {pos: {src: [] for src, _ in EXPERT_SOURCES} for pos in EXPERT_POSITIONS}
    for r in EXPERT_ROWS:
        rank = ws.cell(row=r, column=1).value
        if rank is None:
            continue
        for src, col0 in EXPERT_SOURCES:
            for i, pos in enumerate(EXPERT_POSITIONS):
                val = ws.cell(row=r, column=col0 + i).value
                if val:
                    raw[pos][src].append((int(rank), str(val).strip()))

    result = {}
    for pos in EXPERT_POSITIONS:
        # Build the candidate pool (and team lookup) from the five sources
        # that give full names, before touching the abbreviated Trade/Cut list.
        candidates = {}
        teams = {}
        for src in ["FantasyPros", "ANDY", "JASON", "MIKE", "DraftSharks"]:
            for _rank, name in raw[pos][src]:
                team_match = TEAM_SUFFIX_RE.search(name)
                display = TEAM_SUFFIX_RE.sub("", name).strip()
                key = normalize_name(display)
                candidates.setdefault(key, display)
                if team_match:
                    teams.setdefault(key, team_match.group(1).upper())

        pos_result = {}
        for src, _col0 in EXPERT_SOURCES:
            for rank, token in raw[pos][src]:
                if src == "Trade/Cut":
                    display = resolve_abbrev(pos, token, candidates)
                    if display is None:
                        continue  # couldn't confidently resolve — skip rather than guess
                else:
                    display = TEAM_SUFFIX_RE.sub("", token).strip()
                key = normalize_name(display)
                entry = pos_result.setdefault(key, {"display": display, "team": teams.get(key), "ranks": {}})
                entry["ranks"][src] = rank
        result[pos] = pos_result
    return result


STATS_ROW_RE = re.compile(r'<tr data-row="\d+">.*?</tr>', re.S)
STATS_NAME_MARKER_RE = re.compile(r"[*+]+$")  # PFR appends * (Pro Bowl) / + (All-Pro) to names
STATS_VALID_POS = {"QB", "RB", "WR", "TE"}


def stats_cell(row_html, stat):
    m = re.search(rf'data-stat="{stat}"[^>]*>([^<]*)<', row_html)
    return m.group(1) if m and m.group(1) != "" else None


def load_2025_stats(path):
    """pos -> {normalized_name: {display, team, games, ppg, posRank}}

    Pro-Football-Reference's fantasy-leaders "Share & Export -> Get table as
    XLS" is actually an HTML table (the .xls extension is a PFR quirk, not a
    real Excel file), so this is parsed as HTML rather than via openpyxl.
    """
    content = path.read_text(encoding="utf-8", errors="ignore")
    result = {pos: {} for pos in STATS_VALID_POS}
    for row in STATS_ROW_RE.findall(content):
        pos = stats_cell(row, "fantasy_pos")
        if pos not in STATS_VALID_POS:
            continue
        name = stats_cell(row, "player")
        if not name:
            continue
        name = STATS_NAME_MARKER_RE.sub("", name).strip()
        games = stats_cell(row, "g")
        ppr = stats_cell(row, "fantasy_points_ppr")
        pos_rank = stats_cell(row, "fantasy_rank_pos")
        games = int(games) if games else 0
        ppg = round(float(ppr) / games, 1) if ppr and games else None
        key = normalize_name(name)
        result[pos][key] = {
            "display": name,
            "team": stats_cell(row, "team"),
            "games": games,
            "ppg": ppg,
            "posRank": int(pos_rank) if pos_rank else None,
        }
    return result


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "2026_Dynasty_Startup_Draft_Prep.xlsx"
    stats_src = Path(sys.argv[2]) if len(sys.argv) > 2 else Path.home() / "Downloads" / "sportsref_download.xls"
    wb = openpyxl.load_workbook(src, data_only=True)

    tier_exact, tier_by_last = load_tiers(wb)
    expert = load_expert_consensus(wb)
    expert_by_last = {
        pos: {} for pos in EXPERT_POSITIONS
    }
    for pos, pos_experts in expert.items():
        for key in pos_experts:
            expert_by_last[pos].setdefault(last_name(key), []).append(key)

    ws = wb[RANK_SHEET]
    players = []
    unmatched_tier = []
    consumed_expert_keys = {pos: set() for pos in EXPERT_POSITIONS}

    def attach_expert(player, pos, key):
        pos_experts = expert.get(pos, {})
        entry = pos_experts.get(key)
        used_key = key
        if entry is None:
            candidates = expert_by_last.get(pos, {}).get(last_name(key))
            if candidates and len(candidates) == 1:
                used_key = candidates[0]
                entry = pos_experts.get(used_key)
        if entry:
            ranks = list(entry["ranks"].values())
            player["expertBest"] = min(ranks)
            player["expertWorst"] = max(ranks)
            player["expertAvg"] = round(sum(ranks) / len(ranks), 1)
            player["expertCount"] = len(ranks)
            player["expertRanks"] = entry["ranks"]
            if not player.get("team") and entry.get("team"):
                player["team"] = entry["team"]
            consumed_expert_keys[pos].add(used_key)
        else:
            player["expertBest"] = None
            player["expertWorst"] = None
            player["expertAvg"] = None
            player["expertCount"] = None
            player["expertRanks"] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        rank, name, pos, team, bye, age = row[0], row[1], row[2], row[3], row[4], row[5]
        if rank is None or pos not in VALID_POS or not name:
            continue
        key = normalize_name(name)
        tier_entry = tier_exact.get((pos, key))
        if tier_entry is None:
            candidates = tier_by_last.get((pos, last_name(key)))
            if candidates and len(candidates) == 1:
                tier_entry = next(iter(candidates))
        if tier_entry is None:
            unmatched_tier.append(f"{name} ({pos})")
        player = {
            "rank": int(rank),
            "name": str(name).strip(),
            "pos": pos,
            "team": (team or "").strip(),
            "bye": str(bye).strip() if bye is not None else "",
            "age": age,
            "tier": tier_entry[0] if tier_entry else None,
        }
        attach_expert(player, pos, key)
        players.append(player)

    # Anyone in the expert sheet's top 20 who isn't already in the 472-list —
    # append them so their expert data (and existence) isn't silently lost.
    # Ordered by expert average rank (best first) purely so the append order
    # is sane; where exactly they belong in the real ranking is a manual call.
    new_entries = []
    for pos, pos_experts in expert.items():
        for key, entry in pos_experts.items():
            if key in consumed_expert_keys[pos]:
                continue
            ranks = list(entry["ranks"].values())
            new_entries.append({
                "name": entry["display"],
                "pos": pos,
                "team": entry.get("team") or "",
                "bye": "",
                "age": None,
                "tier": None,
                "expertBest": min(ranks),
                "expertWorst": max(ranks),
                "expertAvg": round(sum(ranks) / len(ranks), 1),
                "expertCount": len(ranks),
                "expertRanks": entry["ranks"],
            })
    new_entries.sort(key=lambda p: p["expertAvg"])
    next_rank = max((p["rank"] for p in players), default=0) + 1
    for p in new_entries:
        p["rank"] = next_rank
        next_rank += 1
        players.append(p)

    # 2025 season stats (games/PPG/season-end position rank), for the
    # Compare Players tool — not currently shown on the main dashboard.
    stats_matched = 0
    if stats_src.exists():
        stats = load_2025_stats(stats_src)
        stats_by_last = {pos: {} for pos in STATS_VALID_POS}
        for pos, pos_stats in stats.items():
            for key in pos_stats:
                stats_by_last[pos].setdefault(last_name(key), []).append(key)

        for player in players:
            pos = player["pos"]
            if pos not in stats:
                continue
            key = normalize_name(player["name"])
            entry = stats[pos].get(key)
            if entry is None:
                candidates = stats_by_last[pos].get(last_name(key))
                if candidates and len(candidates) == 1:
                    entry = stats[pos][candidates[0]]
            if entry:
                player["games2025"] = entry["games"]
                player["ppg2025"] = entry["ppg"]
                player["posRank2025"] = entry["posRank"]
                stats_matched += 1
            else:
                player["games2025"] = None
                player["ppg2025"] = None
                player["posRank2025"] = None
    else:
        print(f"(2025 stats file not found at {stats_src} — skipping, not a hard failure)")
        for player in players:
            player["games2025"] = None
            player["ppg2025"] = None
            player["posRank2025"] = None

    out = Path(__file__).parent / "rankings.json"
    out.write_text(json.dumps(players, indent=0))
    print(f"Wrote {len(players)} players to {out} ({len(new_entries)} newly added from the expert sheet, {stats_matched} matched to 2025 stats)")
    if unmatched_tier:
        print(f"{len(unmatched_tier)} player(s) had no tier match:")
        for u in unmatched_tier:
            print(f"  - {u}")


if __name__ == "__main__":
    main()
