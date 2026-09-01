#!/usr/bin/env python3
"""Rebuilds rankings_redraft.json from the redraft prep workbook:
  - overall rank/name/team/pos from 'Consensus Average' (the 4-source
    consensus big board — the sheet this ranking is meant to go off of)
  - tier from 'BDGE', which is itself a single tiered board (tier breaks are
    "   TIER N    — n PLAYERS" header rows, not per-position tabs like the
    dynasty workbook) — tiers are combined across positions, which is fine
    since the dashboard only ever compares tier numbers within one position
    at a time
  - bye week from BDGE, falling back to Ultimate Draft Guide then Draft
    Sharks for players BDGE doesn't cover

No age (the workbook has no birthdate/age data, and it matters much less
in a one-year redraft league anyway) and no expert-consensus/2025-stats
fields (those exist only to feed the Compare Players tool, which isn't
being built for redraft).

Re-run any time the spreadsheet is updated:

    python3 extract_redraft_rankings.py ~/Downloads/2026_Superflex_Redraft_Rankings.xlsx
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

CONSENSUS_SHEET = "Consensus Average"
TIER_SHEET = "BDGE"
BYE_FALLBACK_SHEETS = ["Ultimate Draft Guide", "Draft Sharks"]  # plain POS column, no rank suffix
VALID_POS = {"QB", "RB", "WR", "TE"}

SUFFIX_RE = re.compile(r"\b(jr|sr|ii|iii|iv|v)\b")
POS_PREFIX_RE = re.compile(r"^([A-Za-z]+)")
TIER_HDR_RE = re.compile(r"TIER\s*(\d+)", re.I)


def normalize_name(name):
    if not name:
        return ""
    n = unicodedata.normalize("NFD", str(name)).encode("ascii", "ignore").decode()
    n = n.lower()
    n = n.replace(".", "").replace("'", "").replace("’", "")
    n = n.replace("-", " ")
    n = SUFFIX_RE.sub("", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def last_name(normalized):
    parts = normalized.split(" ")
    return parts[-1] if parts else ""


def first_names_compatible(a, b):
    """True if two normalized names' first tokens look like the same person
    (exact, or one a prefix of the other — "cam"/"cameron"). Guards the
    last-name fallback below against colliding two different players who
    happen to share a surname (e.g. Bijan Robinson vs Brian Robinson Jr.) —
    sharing a last name alone isn't enough evidence they're the same person."""
    fa = a.split(" ")[0] if a else ""
    fb = b.split(" ")[0] if b else ""
    if not fa or not fb:
        return False
    return fa == fb or fa.startswith(fb) or fb.startswith(fa)


def split_pos(raw):
    """'RB1' -> 'RB'. Returns None if it doesn't start with a known position."""
    m = POS_PREFIX_RE.match(str(raw or ""))
    pos = m.group(1) if m else None
    return pos if pos in VALID_POS else None


def load_tier_and_bye(wb):
    """(pos, normalized_name) -> {"tier": int|None, "bye": str}, plus a
    last-name fallback map, both built from the BDGE sheet (tiered board)."""
    exact = {}
    by_last = {}  # (pos, last) -> set of normalized_name
    ws = wb[TIER_SHEET]
    tier = None
    for row in ws.iter_rows(min_row=5, values_only=True):
        first = row[0]
        if isinstance(first, str) and TIER_HDR_RE.search(first):
            tier = int(TIER_HDR_RE.search(first).group(1))
            continue
        name = row[1]
        if not name:
            continue
        pos = split_pos(row[3])
        if not pos:
            continue
        key = normalize_name(name)
        bye = row[4]
        exact[(pos, key)] = {"tier": tier, "bye": str(bye).strip() if bye not in (None, "—") else ""}
        by_last.setdefault((pos, last_name(key)), set()).add(key)
    return exact, by_last


def load_bye_fallback(wb):
    """(pos, normalized_name) -> bye, from sheets with a plain (non-suffixed)
    POS column, checked in order — first hit wins, doesn't overwrite BDGE."""
    result = {}
    for sheet_name in BYE_FALLBACK_SHEETS:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=5, values_only=True):
            name = row[1]
            if not name:
                continue
            pos = row[3] if row[3] in VALID_POS else None
            if not pos:
                continue
            bye = row[4]
            if bye in (None, "—"):
                continue
            key = (pos, normalize_name(name))
            result.setdefault(key, str(bye).strip())
    return result


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "2026_Superflex_Redraft_Rankings.xlsx"
    wb = openpyxl.load_workbook(src, data_only=True)

    tier_exact, tier_by_last = load_tier_and_bye(wb)
    bye_fallback = load_bye_fallback(wb)

    ws = wb[CONSENSUS_SHEET]
    players = []
    unmatched_tier = 0
    unmatched_bye = 0

    for row in ws.iter_rows(min_row=5, values_only=True):
        rank, name, team = row[0], row[1], row[2]
        pos = split_pos(row[3])
        if rank is None or pos is None or not name:
            continue
        key = normalize_name(name)

        entry = tier_exact.get((pos, key))
        if entry is None:
            candidates = tier_by_last.get((pos, last_name(key)))
            if candidates and len(candidates) == 1:
                only = next(iter(candidates))
                if first_names_compatible(key, only):
                    entry = tier_exact.get((pos, only))

        tier = entry["tier"] if entry else None
        bye = entry["bye"] if entry and entry["bye"] else bye_fallback.get((pos, key), "")

        if tier is None:
            unmatched_tier += 1
        if not bye:
            unmatched_bye += 1

        players.append({
            "rank": int(rank),
            "name": str(name).strip(),
            "pos": pos,
            "team": (team or "").strip(),
            "bye": bye,
            "tier": tier,
        })

    out = Path(__file__).parent / "rankings_redraft.json"
    out.write_text(json.dumps(players, indent=0))
    print(f"Wrote {len(players)} players to {out}")
    print(f"{unmatched_tier} player(s) had no tier match, {unmatched_bye} had no bye match (both just show as blank/'—' in the UI)")


if __name__ == "__main__":
    main()
